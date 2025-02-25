from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import time
import openpyxl

file_path = "C:/Users/Admin/Downloads/Credentials.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

email = sheet.cell(row=2, column=1).value  # First column (Email)
password = sheet.cell(row=2, column=2).value  # Second column (Password)

# Setup Chrome WebDriver
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
#driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver = webdriver.Chrome(service=Service("C:/Users/Admin/Downloads/chromedriver-win64/chromedriver-win64/chromedriver.exe"), options=options)


# Open Google
driver.get("https://www.hpsmart.com")

# Wait up to 10 seconds for the button to be clickable
wait = WebDriverWait(driver, 10)

accept_button = wait.until(EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler")))
accept_button.click()

create_account_button = wait.until(EC.element_to_be_clickable((By.ID,"create-account-button")))
create_account_button.click()

First_Name = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='firstName']")))
First_Name.click()
First_Name.send_keys("Sowmya")

Last_Name = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='lastName']")))
Last_Name.click()
Last_Name.send_keys("Guruswamy")

email_Xpath = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='email']")))
email_Xpath.click()
email_Xpath.send_keys(email)

password_Xpath = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@id='password']")))
password_Xpath.click()
password_Xpath.send_keys(password)

Submit_Button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@id='sign-up-submit']")))
Submit_Button.click()

print("Successfully created account")

time.sleep(2)  # Wait for the page to load

input("Perform puzzle")
# Close the browser

#post account creation

#Getting_Most = wait.until(EC._element_if_visible((By.NAME, "Getting the most out of your account")))
#Continue_ID = driver.find_element(By.ID,"full-screen-consent-form-footer-button-continue")

# Check if the element is displayed and enabled
try:
    # Find the element
    Continue_ID = driver.find_element(By.ID, 'continue-button-id')
    driver.get_screenshot_as_png()
    # Check if the element is displayed and enabled
    if Continue_ID.is_displayed():
        if Continue_ID.is_enabled():
            Continue_ID.click()
            print("Element clicked!")
        else:
            print("Element is displayed but not enabled.")
    else:
        print("Element is not displayed.")
except Exception as e:
    print(f"Error: {e}")

driver.quit()
